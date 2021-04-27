import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QToolTip, QMainWindow, QProgressBar
from PyQt5.QtGui import QFont
from PyQt5.QtCore import QThread, pyqtSignal
import urllib.request
from bs4 import BeautifulSoup
import urllib.parse
from selenium import webdriver
import time
import pandas as pd
import openpyxl
import os
import math
import progress


class ThreadClass(QThread):

    percentage = pyqtSignal(int)

    def __init__(self, parent=None):
        super(ThreadClass, self).__init__(parent)

    def run(self):

        wb = openpyxl.load_workbook("base_excel.xlsx")
        browser = self.start_webdriver()

        codes = self.get_all_code()

        for i, code in enumerate(codes):

            number_code = code["code"]
            name_code = code["name"]

            sheet = self.create_sheet(wb, "Sheet1", name_code)
            self.fill_sheet(number_code, browser, sheet)

            progress_step = math.floor(((i+1) / len(codes) * 100))
            self.percentage.emit(progress_step)

            print(i)

        if os.path.exists("Naver_Finance_Data.xlsx"):
            os.remove("Naver_Finance_Data.xlsx")

        wb.save("Naver_Finance_Data.xlsx")

    def get_all_code(self):

        df = pd.read_excel('종목코드.xlsx')
        code_col = df['code']
        name_col = df['name']

        code_list = []

        for i in range(len(df)):
            code = code_col[i]
            name = name_col[i]

            for i in range(6 - len(str(code))):
                code = '0' + str(code)

            code_list.append({"code": str(code), "name": name})

        return code_list

    def start_webdriver(self):

        option = webdriver.ChromeOptions()
        option.add_argument('headless')
        browser = webdriver.Chrome(options=option)

        return browser

    def create_sheet(self, wb, source, sheet_name):

        source = wb[source]
        target = wb.copy_worksheet(source)

        print(type(target))

        target.title = sheet_name
        wb.name = target

        return target

    def fill_sheet(self, code, browser, sheet):

        url_annual = 'https://m.stock.naver.com/item/main.nhn#/stocks/{}/annual'.format(
            code)

        browser.get(url_annual)

        time.sleep(2)

        net_profit = browser.find_elements_by_xpath(
            '/html/body/div[2]/div[1]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[2]/td/span')

        periods = browser.find_elements_by_xpath(
            '/html/body/div[2]/div[1]/div[2]/div[1]/div[2]/div/div/table/thead/tr/th/span/span[1]')

        actual_profit = browser.find_elements_by_xpath(
            '/html/body/div[2]/div[1]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[3]/td/span')

        try:

            # 연간 기간
            sheet['C6'] = periods[0].text[:7]
            sheet['D6'] = periods[1].text[:7]
            sheet['E6'] = periods[2].text[:7]
            sheet['F6'] = periods[3].text[:7]

            # 연간 영업이익
            sheet['C7'] = net_profit[0].text
            sheet['D7'] = net_profit[1].text
            sheet['E7'] = net_profit[2].text
            sheet['F7'] = net_profit[3].text

            # 연간 영업이익 성장률
            # ( current - previous ) / previous
            # sheet['C8'] = 'N/A'

            if sheet['D7'].value != '' and sheet['C7'].value != '' and sheet['C7'].value != '0':
                sheet['D8'] = (int(sheet['D7'].value.replace(',', '')) -
                               int(sheet['C7'].value.replace(',', ''))) / abs(int(sheet['C7'].value.replace(',', '')))

            if sheet['E7'].value != '' and sheet['D7'].value != '' and sheet['D7'].value != '0':
                sheet['E8'] = (int(sheet['E7'].value.replace(',', '')) -
                               int(sheet['D7'].value.replace(',', ''))) / abs(int(sheet['D7'].value.replace(',', '')))

            if sheet['E7'].value != '' and sheet['F7'].value != '' and sheet['E7'].value != '0':
                sheet['F8'] = (int(sheet['F7'].value.replace(',', '')) -
                               int(sheet['E7'].value.replace(',', ''))) / abs(int(sheet['E7'].value.replace(',', '')))

            # 연간 당기순이익
            sheet['C9'] = actual_profit[0].text
            sheet['D9'] = actual_profit[1].text
            sheet['E9'] = actual_profit[2].text
            sheet['F9'] = actual_profit[3].text

            # 연간 당기순이익 성장률
            # sheet['C10'] = 'N/A'
            if sheet['D9'].value != '' and sheet['C9'].value != '' and sheet['C9'].value != '0':
                sheet['D10'] = (int(sheet['D9'].value.replace(',', '')) -
                                int(sheet['C9'].value.replace(',', ''))) / abs(int(sheet['C9'].value.replace(',', '')))

            if sheet['E9'].value != '' and sheet['D9'].value != '' and sheet['D9'].value != '0':
                sheet['E10'] = (int(sheet['E9'].value.replace(',', '')) -
                                int(sheet['D9'].value.replace(',', ''))) / abs(int(sheet['D9'].value.replace(',', '')))

            if sheet['F9'].value != '' and sheet['E9'].value != '' and sheet['E9'].value != '0':
                sheet['F10'] = (int(sheet['F9'].value.replace(',', '')) -
                                int(sheet['E9'].value.replace(',', ''))) / abs(int(sheet['E9'].value.replace(',', '')))

        except:
            pass

        url_quarter = 'https://m.stock.naver.com/item/main.nhn#/stocks/{}/quarter'.format(
            code)

        browser.get(url_quarter)

        time.sleep(2)

        net_profit_quarter = browser.find_elements_by_xpath(
            '/html/body/div[2]/div[1]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[2]/td/span')

        periods_quarter = browser.find_elements_by_xpath(
            '/html/body/div[2]/div[1]/div[2]/div[1]/div[2]/div/div/table/thead/tr/th/span/span[1]')

        actual_profit_quarter = browser.find_elements_by_xpath(
            '/html/body/div[2]/div[1]/div[2]/div[1]/div[2]/div/div/table/tbody/tr[3]/td/span')

        try:

            # 분기 기간
            sheet['C12'] = periods_quarter[2].text[:7]
            sheet['D12'] = periods_quarter[3].text[:7]
            sheet['E12'] = periods_quarter[4].text[:7]
            sheet['F12'] = periods_quarter[5].text[:7]

            # 분기 영업이익
            sheet['C13'] = net_profit_quarter[2].text
            sheet['D13'] = net_profit_quarter[3].text
            sheet['E13'] = net_profit_quarter[4].text
            sheet['F13'] = net_profit_quarter[5].text

            # 분기 영업이익 성장률
            # sheet['C14'] = 'N/A'
            if sheet['D13'].value != '' and sheet['C13'].value != '' and sheet['C13'].value != '0':
                sheet['D14'] = (int(sheet['D13'].value.replace(',', '')) -
                                int(sheet['C13'].value.replace(',', ''))) / abs(int(sheet['C13'].value.replace(',', '')))

            if sheet['E13'].value != '' and sheet['D13'].value != '' and sheet['D13'].value != '0':
                sheet['E14'] = (int(sheet['E13'].value.replace(',', '')) -
                                int(sheet['D13'].value.replace(',', ''))) / abs(int(sheet['D13'].value.replace(',', '')))

            if sheet['F13'].value != '' and sheet['E13'].value != '' and sheet['E13'].value != '0':
                sheet['F14'] = (int(sheet['F13'].value.replace(',', '')) -
                                int(sheet['E13'].value.replace(',', ''))) / abs(int(sheet['E13'].value.replace(',', '')))

            # 분기 당기순이익
            sheet['C15'] = actual_profit_quarter[2].text
            sheet['D15'] = actual_profit_quarter[3].text
            sheet['E15'] = actual_profit_quarter[4].text
            sheet['F15'] = actual_profit_quarter[5].text

            # 분기 당기순이익 성장률
            # sheet['C16'] = 'N/A'

            if sheet['D15'].value != '' and sheet['C15'].value != '' and sheet['C15'].value != '0':
                sheet['D16'] = (int(sheet['D15'].value.replace(',', '')) -
                                int(sheet['C15'].value.replace(',', ''))) / abs(int(sheet['C15'].value.replace(',', '')))

            if sheet['D15'].value != '' and sheet['E15'].value != '' and sheet['D15'].value != '0':
                sheet['E16'] = (int(sheet['E15'].value.replace(',', '')) -
                                int(sheet['D15'].value.replace(',', ''))) / abs(int(sheet['D15'].value.replace(',', '')))

            if sheet['F15'].value != '' and sheet['E15'].value != '' and sheet['E15'].value != '0':
                sheet['F16'] = (int(sheet['F15'].value.replace(',', '')) -
                                int(sheet['E15'].value.replace(',', ''))) / abs(int(sheet['E15'].value.replace(',', '')))

        except:
            pass


class MyApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        self.btn = QPushButton("Crawl Data", self)
        self.btn.move(150, 50)
        self.btn.resize(self.btn.sizeHint())
        self.btn.clicked.connect(self.startProgressBar)

        self.pbar = QProgressBar(self)
        self.pbar.setGeometry(80, 100, 280, 25)
        self.statusBar().showMessage('Ready')

        self.setWindowTitle('네이버 증권 크롤러')
        self.move(300, 300)
        self.resize(400, 200)
        self.show()

    def startProgressBar(self):
        self.thread = ThreadClass()
        self.thread.percentage.connect(self.updateProgressBar)
        self.thread.start()

    def updateProgressBar(self, val):
        self.pbar.setValue(val)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
